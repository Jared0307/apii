import requests
from collections import defaultdict
import openpyxl
from openpyxl.styles import Alignment, Font

# Definir las URL base para cada ruta de la API
base_url = 'http://localhost:3001'  # Reemplaza esto con la URL de tu API
endpoints = [
    '/categorias', 
    '/estados',
    '/marcas',
    '/productos',
    '/roles', 
    '/usuarios',
    '/ventas'
    # Agrega aquí los endpoints que necesites para las otras consultas
]
cantidad_minima = 10  # Define el umbral mínimo para reabastecer productos

# Función para hacer peticiones a la API y obtener datos
def get_data(endpoint):
    url = f"{base_url}{endpoint}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error al obtener datos de {endpoint}: {response.status_code}")
        return None

# Ventas Totales
ventas_data = get_data('/ventas')

# Obtener los datos de /ventas
response_ventas = requests.get(f"{base_url}/ventas")

if response_ventas.status_code == 200:
    ventas_data = response_ventas.json()

    # Obtener los datos de /productos
    response_productos = requests.get(f"{base_url}/productos")

    if response_productos.status_code == 200:
        productos_data = response_productos.json()
        precios_productos = {producto['id_producto']: float(producto['precio']) for producto in productos_data}
        total_ingreso = sum(precios_productos[venta['id_producto']] * venta['cantidad'] for venta in ventas_data if venta['id_producto'] in precios_productos)
    else:
        print(f"Error al obtener datos de productos: {response_productos.status_code}")
else:
    print(f"Error al obtener datos de ventas: {response_ventas.status_code}")

# Obtener las cantidades totales vendidas por categoría
ventas_por_categoria = defaultdict(int)

for venta in ventas_data:
    id_producto = venta['id_producto']
    cantidad = venta['cantidad']
    # Encuentra el producto en los datos de productos
    for producto in productos_data:
        if producto['id_producto'] == id_producto:
            id_categoria = producto['id_categoria']
            ventas_por_categoria[id_categoria] += cantidad

# Encuentra la categoría más vendida
categoria_mas_vendida_id = max(ventas_por_categoria, key=ventas_por_categoria.get)

# Obtener el nombre de la categoría más vendida usando la API de categorías
categoria_mas_vendida = get_data(f'/categorias/{categoria_mas_vendida_id}')

# Crear un diccionario para mapear los precios por id_producto
precios_productos = {producto['id_producto']: float(producto['precio']) for producto in productos_data}

# Obtener un diccionario para mapear los nombres de usuario por su ID
usuarios_data = get_data('/usuarios')
usuarios_nombres = {usuario['id_usuario']: f"{usuario['nombre_usuario']} {usuario['apellido_usuario']}" for usuario in usuarios_data}

# Calculamos el gasto total de cada usuario
usuarios_gastos = defaultdict(float)

for venta in ventas_data:
    id_usuario = venta['id_usuario']
    id_producto = venta['id_producto']
    cantidad = venta['cantidad']
    precio = precios_productos.get(id_producto, 0)  # Obtener el precio del producto o 0 si no se encuentra
    usuarios_gastos[id_usuario] += precio * cantidad

# Encontrar el usuario que más ha gastado
usuario_max_gasto_id = max(usuarios_gastos, key=usuarios_gastos.get)
gasto_maximo = usuarios_gastos[usuario_max_gasto_id]

# Obtener el nombre del usuario que más ha gastado
usuario_max_gasto_nombre = usuarios_nombres.get(usuario_max_gasto_id, "Nombre no encontrado")

# Usuarios y Comportamiento de Compra
from collections import defaultdict

# Para calcular el ingreso total:
ingreso_total = sum(
    float(next(producto['precio'] for producto in productos_data if producto['id_producto'] == venta['id_producto']) or 0) * venta['cantidad']
    for venta in ventas_data
)

# Encontrar el producto más vendido
productos_vendidos = defaultdict(int)
for venta in ventas_data:
    productos_vendidos[venta['id_producto']] += venta['cantidad']
producto_mas_vendido_id = max(productos_vendidos, key=productos_vendidos.get)

# Encontrar la categoría más vendida
categorias_vendidas = defaultdict(int)
for venta in ventas_data:
    producto = next((p for p in productos_data if p['id_producto'] == venta['id_producto']), None)
    if producto:
        categorias_vendidas[producto['id_categoria']] += venta['cantidad']
categoria_mas_vendida_id = max(categorias_vendidas, key=categorias_vendidas.get)

# Para calcular los gastos de cada usuario:
usuarios_gastos = {}
for venta in ventas_data:
    id_producto = venta['id_producto']
    id_usuario = venta['id_usuario']
    
    # Buscamos el precio del producto en la lista de productos
    precio_producto = next((producto['precio'] for producto in productos_data if producto['id_producto'] == id_producto), 0)
    
    # Calculamos el gasto y lo agregamos al usuario correspondiente
    usuarios_gastos[id_usuario] = usuarios_gastos.get(id_usuario, 0) + float(precio_producto) * venta['cantidad']

# Cálculo del promedio de gasto por usuario
promedio_gasto = sum(usuarios_gastos.values()) / len(usuarios_gastos)

# Cálculo de cuántos productos compra un usuario en promedio
productos_por_usuario = defaultdict(int)
for venta in ventas_data:
    productos_por_usuario[venta['id_usuario']] += venta['cantidad']
promedio_productos_por_usuario = sum(productos_por_usuario.values()) / len(productos_por_usuario)

# Cálculo de productos en stock
productos_en_stock = sum(producto['stock'] for producto in productos_data)

# Encontrar el producto con menos stock
producto_menos_stock_id = min(productos_data, key=lambda x: x['stock'])['id_producto']

# Contar productos que necesitan ser reabastecidos (stock bajo)
productos_bajo_stock = sum(1 for producto in productos_data if producto['stock'] < 10)
usuario_max_gasto_nombre = next(usuario['nombre_usuario'] for usuario in usuarios_data if usuario['id_usuario'] == usuario_max_gasto_id)

# Encontrar el producto más caro
producto_mas_caro = max(productos_data, key=lambda x: float(x['precio']))

# Encontrar el producto más barato
producto_mas_barato = min(productos_data, key=lambda x: float(x['precio']))

# Cálculo del promedio de precios de los productos
promedio_precios = sum(float(producto['precio']) for producto in productos_data) / len(productos_data)

# Crear un nuevo libro de trabajo
workbook = openpyxl.Workbook()
sheet = workbook.active

# Preguntas y respuestas
preguntas = [
    "PREGUNTAS",
    "Ventas:",
    "",
    "¿Cuánto es el ingreso total de todas las ventas?",
    "¿Cuál es el producto más vendido?",
    "¿Cuál es la categoría más vendida?",
    "",
    "Usuarios:",
    "",
    "¿Cuál es el usuario que más ha gastado?",
    "¿Cuál es el promedio de gasto por usuario?",
    "Cuántos productos compra un usuario en promedio?",
    "",
    "Inventario:",
    "",
    "¿Cuántos productos están actualmente en stock?",
    "¿Cuál es el producto con menos stock?",
    "¿Cuántos productos necesitan ser reabastecidos (stock bajo = 10)?",
    "",
    "Productos:",
    "",
    "¿Cuál es el producto más caro?",
    "¿Cuál es el producto más barato?",
    "¿Cuál es el promedio de precios de los productos"
]

respuestas = [
    "RESPUESTAS",
    "Ventas:",
    "",
    f"${ingreso_total}",
    f"{next(p['nombre_producto'] for p in productos_data if p['id_producto'] == producto_mas_vendido_id)}",
    f"{categoria_mas_vendida_id}",
    "",
    "Usuarios:",
    "",
    f"{usuario_max_gasto_nombre}",
    f"${promedio_gasto}",
    f"{promedio_productos_por_usuario}",
    "",
    "Inventario:",
    "",
    f"{productos_en_stock}",
    f"{next(p['nombre_producto'] for p in productos_data if p['id_producto'] == producto_menos_stock_id)}",
    f"{productos_bajo_stock}",
    "",
    "Productos:",
    "",
    f"{producto_mas_caro['nombre_producto']}",
    f"{producto_mas_barato['nombre_producto']}",
    f"${promedio_precios}"
    # Agregar el resto de respuestas utilizando las variables correspondientes
]

# Escribir en el archivo Excel
for idx, pregunta in enumerate(preguntas):
    sheet.cell(row=idx + 1, column=1, value=pregunta)
    if idx == 0:
        sheet.cell(row=idx + 1, column=1).font = Font(b=True)  # Aplicar negrita solo en la primera celda de la primera columna

for idx, respuesta in enumerate(respuestas):
    sheet.cell(row=idx + 1, column=2, value=respuesta)
    if idx == 0:
        sheet.cell(row=idx + 1, column=2).font = Font(b=True)  # Aplicar negrita solo en la primera celda de la segunda columna

# Aplicar estilos adicionales a filas específicas
filas_especiales = [2, 8, 14, 20]
for fila in filas_especiales:
    sheet.cell(row=fila, column=1).font = Font(size=13, italic=True, name='Arial')  # Aplicar estilo a la primera columna
    sheet.cell(row=fila, column=2).font = Font(size=13, italic=True, name='Arial')  # Aplicar estilo a la segunda columna
    sheet.cell(row=fila, column=1).alignment = Alignment(horizontal='left')  # Alinear a la izquierda la primera columna
    sheet.cell(row=fila, column=2).alignment = Alignment(horizontal='right')  # Alinear a la izquierda la segunda columna

# Aplicar estilos adicionales a filas específicas
filas_especiales2 = [4, 5, 6, 10, 11, 12, 16, 17, 18, 22, 23, 24]
for fila2 in filas_especiales2:
    sheet.cell(row=fila2, column=1).font = Font(size=15, italic=True, name='Arial Narrow')  # Aplicar estilo a la primera columna
    sheet.cell(row=fila2, column=2).font = Font(size=15, italic=True, name='Arial Narrow')  # Aplicar estilo a la segunda columna
    sheet.cell(row=fila2, column=1).alignment = Alignment(horizontal='right')  # Alinear a la izquierda la primera columna
    sheet.cell(row=fila2, column=2).alignment = Alignment(horizontal='left')  # Alinear a la izquierda la segunda columna

# Aplicar estilos adicionales a filas específicas
filas_especiales3 = [1]
for fila3 in filas_especiales3:
    sheet.cell(row=fila3, column=1).font = Font(size=15, italic=True, name='Cambria')  # Aplicar estilo a la primera columna
    sheet.cell(row=fila3, column=2).font = Font(size=15, italic=True, name='Cambria')  # Aplicar estilo a la segunda columna
    sheet.cell(row=fila3, column=1).alignment = Alignment(horizontal='center')  # Alinear a la izquierda la primera columna
    sheet.cell(row=fila3, column=2).alignment = Alignment(horizontal='center')  # Alinear a la izquierda la segunda columna
    
# Guardar el archivo
workbook.save('Preguntas.xlsx')


jared = "1"
        